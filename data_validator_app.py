#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import streamlit as st
import pandas as pd
import json
import re
from io import BytesIO
import chardet
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import locale
from openai import OpenAI

# ✅ 한글 가나다 정렬을 위한 로케일 설정
locale.setlocale(locale.LC_ALL, '')

# ✅ 기준 디렉토리 설정
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
META_DIR = os.path.join(BASE_DIR, "meta_dicts_final_clean")

# ✅ 메타 사전 로딩 (키 공백 제거)
def load_meta_dict(standard):
    path = os.path.join(META_DIR, f"{standard}.json")
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        original_meta = json.load(f)
    return {k.strip().replace(" ", ""): v for k, v in original_meta.items()}

# ✅ GPT 기반 정규식 생성 함수 (OpenAI SDK v1.x 방식)
def generate_regex_from_description(description,expression, column_name):
    client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])

    prompt = f"""
당신은 공공데이터 형식 검사를 위한 정규식을 생성하는 전문가입니다.

다음은 공공데이터 표준 항목에 대한 정보입니다.

컬럼명: {column_name}
설명: {description}
표현형식(예시): {expression}

위 정보를 바탕으로, 이 컬럼의 유효성 검사를 위한 정규식을 생성해주세요.

📌 작성 기준:
- 설명 및 예시에서 유추 가능한 형식을 충실히 반영해주세요.
- 값이 단일일 수도 있고, 복수일 경우 특정 구분자(예: '+', ',', '~')로 연결될 수 있습니다.
- 한글, 숫자, 특수기호, 공백 등이 포함될 수 있으므로, 예시를 기반으로 판단해주세요.
- 지나치게 엄격하지 않으면서, 잘못된 형식을 걸러낼 수 있는 범용 정규식을 생성해주세요.

⚠️ 반드시 정규식만 한 줄로 출력해주세요. 따옴표나 설명 없이, 정규식만 주세요.
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        st.warning(f"❗ 정규식 생성 실패: {e}")
        return None

def validate_cell(val, col, meta, row_data):
    errors = []
    val_raw = str(val).strip()
    val_clean = val_raw.upper()

    meta_col = meta.get(col)
    if not meta_col:
        return errors

    required = meta_col.get("필수여부") == "필수"
    조건부 = meta_col.get("조건부필수")

    if val_clean in ["", "NAN", "NA"]:
        if required:
            errors.append("필수값 누락")
        elif 조건부:
            기준필드, 기준값들 = list(조건부.items())[0]
            기준값 = str(row_data.get(기준필드, "")).strip().upper()
            if 기준값 in [v.strip().upper() for v in 기준값들]:
                errors.append("조건부 필수 누락")
        return errors

    # ✅ 정규식 → 허용값 → GPT(description) 순서로 검증
    regex = meta_col.get("정규식")
    allowed = meta_col.get("허용값")
    description = meta_col.get("설명")
    expression = meta_col.get("표현형식")

    if regex:
        try:
            if not re.fullmatch(regex, val_raw):
                errors.append("형식 오류")
        except Exception as e:
            errors.append(f"정규식 오류 ({e})")
    elif allowed:
        allowed_clean = [v.strip().upper() for v in allowed]
        if val_clean not in allowed_clean:
            errors.append("허용값 오류")
    elif description:
        regex = generate_regex_from_description(description, expression, col)
        meta_col["정규식"] = regex  # 캐싱
        try:
            if not re.fullmatch(regex, val_raw):
                errors.append("형식 오류(GPT)")
        except Exception as e:
            errors.append(f"GPT 정규식 오류 ({e})")

    return errors


# ✅ 전체 검증 실행 함수
def run_meta_validation(df, meta):
    error_cells = []
    for i, row in df.iterrows():
        for col in df.columns:
            val = row[col]
            row_data = row.to_dict()
            errs = validate_cell(val, col, meta, row_data)
            if errs:
                error_cells.append((i+2, col, ", ".join(errs)))
    return error_cells

# ✅ 오류 셀 표시된 엑셀 생성
def generate_excel_with_errors(df, error_cells):
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    col_idx = {col: i+1 for i, col in enumerate(df.columns)}
    for row, col, _ in error_cells:
        ws.cell(row=row, column=col_idx[col]).fill = yellow
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# ✅ Streamlit 앱 실행
def data_validator_app():
    st.title("📑 공공데이터 정밀 검증기 (GPT 자동 정규식 생성 포함)")

    uploaded_file = st.file_uploader("📂 CSV 파일을 업로드하세요", type=["csv"])

    if not os.path.exists(META_DIR):
        st.error("❌ meta_dicts_final_clean 폴더가 존재하지 않습니다.")
        st.stop()

    meta_files = [f.replace(".json", "") for f in os.listdir(META_DIR) if f.endswith(".json")]
    meta_files_sorted = sorted(meta_files, key=locale.strxfrm)

    standard = st.selectbox("검증 기준 표준을 선택하세요", options=meta_files_sorted)

    if uploaded_file and standard:
        try:
            raw_bytes = uploaded_file.read()
            encoding = chardet.detect(raw_bytes)['encoding'] or 'utf-8'
            df = pd.read_csv(BytesIO(raw_bytes), encoding=encoding, dtype=str).fillna("")

            df.columns = [col.strip().replace(" ", "") for col in df.columns]

            st.success(f"✅ 파일 업로드 성공 (인코딩: {encoding})")
            st.dataframe(df)

            meta = load_meta_dict(standard)
            if not meta:
                st.error("❌ 메타 정보를 불러올 수 없습니다.")
                return

            if st.button("🔍 정밀 검증 실행"):
                error_cells = run_meta_validation(df, meta)
                st.subheader("📋 검증 결과 미리보기")

                preview_df = df.copy()
                for row, col, msg in error_cells:
                    preview_df.at[row - 2, col] += f" ⚠️ ({msg})"

                st.dataframe(preview_df, use_container_width=True)

                excel_with_errors = generate_excel_with_errors(df, error_cells)
                st.download_button(
                    label="📥 오류 표시된 엑셀 다운로드",
                    data=excel_with_errors.getvalue(),
                    file_name="검증결과_정밀표시.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"❌ 파일 처리 오류: {e}")

